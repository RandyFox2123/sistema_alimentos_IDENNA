from django.shortcuts import render, get_object_or_404
from django.http import HttpResponseRedirect, HttpResponse, Http404
from django.urls import reverse
from django.core.paginator import Paginator
from django.views.decorators.http import require_safe, require_http_methods, require_POST
from openpyxl import Workbook
from  . models import Almacen, Alimento, Medida
import os
import logging
from django.conf import settings
from decimal import Decimal


logger = logging.getLogger(__name__)

error_no_encontrado = 'Lo sentimos, el recurso que solicitas no fue encontrado'
error_inesperado = 'Lo sentimos, pero ha ocurrido un error inesperado en el proceso'
metodo_no_permitido = 'Lo sentimos, hubo un error, el método no está permitido, no se puede proceder por seguridad'

class error_contexto(Exception):
    pass

def manejo_errores(vista_recibida):
    """
    Decorador para manejar errores en las vistas.
    """
    def wrapper(request, *args, **kwargs):
        mensaje_error = None
        try:
            return vista_recibida(request, *args, **kwargs)

        except TypeError as e:
            mensaje_error = 'Lo sentimos, hubo un error interno en el proceso'
            logger.error(f'!!! ERROR OBTENIDO !!!:\n\n{e}')
            return render(request, 'error.html', {'error_obtenido': mensaje_error})

        except Http404 as e:
            mensaje_error = error_no_encontrado
            logger.error(f'!!! ERROR OBTENIDO !!!:\n\n{e}')
            return render(request, 'error.html', {'error_obtenido': mensaje_error})

        except Exception as error_obtenido:
            if isinstance(error_obtenido, error_contexto):
                mensaje_error = str(error_obtenido)
                logger.error(mensaje_error)
            else:
                mensaje_error = error_inesperado
                logger.error(f'{error_obtenido}')
            return render(request, 'error.html', {'error_obtenido': mensaje_error})

    return wrapper



"""===> VISTAS GENERALES <==="""
@manejo_errores
def index(request):
    return render(request, 'index.html')


@manejo_errores
@require_safe
def panel_principal(request):
    almacenes_activos = Almacen.objects.filter(estado=True)
    filtro_desc_alimento = request.GET.get('filtro_desc_alimento', '').strip()
    filtro_almacen_id = request.GET.get('filtro_almacen', '').strip()
    page_num = request.GET.get('page', '')
    alimentos_qs = Alimento.objects.filter(almacen_perteneciente__estado=True)

    if filtro_desc_alimento:
        alimentos_qs = alimentos_qs.filter(desc_alimento__icontains=filtro_desc_alimento).order_by('-id_alimento')

    if filtro_almacen_id.isdigit():
        alimentos_qs = alimentos_qs.filter(almacen_perteneciente_id=int(filtro_almacen_id)).order_by('-id_alimento')

    paginator = Paginator(alimentos_qs.order_by('-id_alimento'), 3)
    page_obj = paginator.get_page(page_num)

    for alimento in page_obj:
        if alimento.imagen_alimento and alimento.imagen_alimento.name:
            ruta_imagen = os.path.join(settings.MEDIA_ROOT, alimento.imagen_alimento.name)
            alimento.imagen_existe = os.path.isfile(ruta_imagen)
        else:
            alimento.imagen_existe = False

    query_params = []
    if filtro_desc_alimento:
        query_params.append(f"filtro_desc_alimento={filtro_desc_alimento}")
    if filtro_almacen_id:
        query_params.append(f"filtro_almacen={filtro_almacen_id}")
    if page_num:
        query_params.append(f"page={page_num}")

    url_con_filtros = reverse('panel_principal')
    if query_params:
        url_con_filtros += '?' + '&'.join(query_params)

    request.session['url_panel_principal'] = url_con_filtros

    contexto = {
        'page_obj': page_obj,
        'almacenes_activos': almacenes_activos,
        'filtro_desc_alimento': filtro_desc_alimento,
        'filtro_almacen_id': filtro_almacen_id,
    }
    return render(request, 'panel_principal.html', contexto)




"""===> VISTAS ALIMENTOS <==="""
@manejo_errores
@require_http_methods(["GET", "POST"])
def registrar_alimento(request):
    almacenes_activos = Almacen.objects.filter(estado=True)
    todas_medidas = Medida.objects.all()
    
    if request.method == 'POST':
        desc_alimento = request.POST.get('desc_alimento', '').strip()
        almacen_id = request.POST.get('almacen_perteneciente')
        imagen = request.FILES.get('imagen_alimento')
        cantidad_str = request.POST.get('cantidad', '0').strip()
        
        
        medida_recibida = request.POST.get('medida', None)
        if medida_recibida != None and medida_recibida.isdigit(): 
            medida_recibida = int(medida_recibida)
            if medida_recibida < 0: error_contexto("Se ha recibido un tipo de cantidad invalido")
            else:medida_obj = Medida.objects.filter(pk=int(medida_recibida)).first()
             
        else:
            error_contexto("Se ha recibido un tipo de cantidad invalido")
             

        if Alimento.objects.filter(desc_alimento__iexact=desc_alimento).exists():
            raise error_contexto("Ya existe un alimento con esta descripcion.")

        if not desc_alimento:
            raise error_contexto("La descripcion es obligatoria.")

        try:
            cantidad = float(cantidad_str.replace(',', '.'))
            #return HttpResponse(f"Cantidad:{cantidad} tipo dato:{type(cantidad)}")
        
            if cantidad < 0 or cantidad > 50000:
                raise error_contexto("La cantidad debe ser un numero mayor o igual a 0 o maximo 50000")
            
        except ValueError:
            raise error_contexto("La cantidad debe ser un numero valido.")

        almacen_obj = None
        if almacen_id and almacen_id.isdigit():
            almacen_obj = Almacen.objects.filter(pk=int(almacen_id), estado=True).first()
            
        #Esta es la parte donde se crea el alimento
        nuevo_alimento = Alimento(
            desc_alimento=desc_alimento, 
            almacen_perteneciente=almacen_obj, 
            cantidad=cantidad, 
            medida_alimento=medida_obj,
        )
        if imagen: nuevo_alimento.imagen_alimento = imagen
        nuevo_alimento.save()

        url_redireccion = request.session.get('url_panel_principal', reverse('panel_principal'))
        return HttpResponseRedirect(url_redireccion)

    return render(request, 'registrar_alimento.html', {'almacenes_activos': almacenes_activos, "medidas":todas_medidas})


@manejo_errores
@require_http_methods(["GET", "POST"])
def editar_alimento(request, id_alimento):
    alimento_obj = get_object_or_404(Alimento, pk=id_alimento)
    almacenes_activos = Almacen.objects.filter(estado=True)
    medidas = Medida.objects.all()

    if request.method == 'POST':
        desc_alimento = request.POST.get('desc_alimento', '').strip()
        almacen_id = request.POST.get('almacen_perteneciente')
        nueva_imagen = request.FILES.get('imagen_alimento')
        cantidad_str = request.POST.get('cantidad', '').strip()
        desc_alimento_vieja = request.POST.get('desc_alimento_vieja', '').strip()

        if desc_alimento_vieja != desc_alimento:
            if Alimento.objects.filter(desc_alimento__iexact=desc_alimento).exists():
                raise error_contexto("Ya existe un alimento con esta descripcion.")
        else:
            desc_alimento = desc_alimento_vieja

        if not desc_alimento:
            raise error_contexto("La descripción es obligatoria.")

        if cantidad_str == '':
            raise error_contexto("La cantidad es obligatoria.")

        try:
            cantidad = float(cantidad_str.replace(',', '.'))
            if cantidad < 0 or cantidad > 50000:
                raise error_contexto("La cantidad debe ser un numero mayor o igual a 0 o maximo 50000")
        except ValueError:
            raise error_contexto("La cantidad debe ser un número válido.")

        almacen_obj = None
        if almacen_id and almacen_id.isdigit():
            almacen_obj = Almacen.objects.filter(pk=int(almacen_id), estado=True).first()

        if nueva_imagen:
            if alimento_obj.imagen_alimento and os.path.isfile(alimento_obj.imagen_alimento.path):
                os.remove(alimento_obj.imagen_alimento.path)
            alimento_obj.imagen_alimento = nueva_imagen

        alimento_obj.desc_alimento = desc_alimento
        alimento_obj.almacen_perteneciente = almacen_obj
        alimento_obj.cantidad = cantidad
        alimento_obj.save()

        url_redireccion = request.session.get('url_panel_principal', reverse('panel_principal'))
        return HttpResponseRedirect(url_redireccion)

    contexto = {
        'alimento_obj': alimento_obj,
        'almacenes_activos': almacenes_activos,
        'medidas': medidas
    }
    return render(request, 'editar_alimento.html', contexto)


@manejo_errores
@require_POST
def borrar_alimento(request, id_alimento):
    alimento_obj = get_object_or_404(Alimento, pk=id_alimento)

    if alimento_obj.imagen_alimento and os.path.isfile(alimento_obj.imagen_alimento.path):
        os.remove(alimento_obj.imagen_alimento.path)

    alimento_obj.delete()

    url_redireccion = request.session.get('url_panel_principal', reverse('panel_principal'))
    return HttpResponseRedirect(url_redireccion)


@manejo_errores
@require_POST
def sumar_cantidad_alimento(request, id_alimento):
    alimento = get_object_or_404(Alimento, pk=id_alimento)

    try:
        cantidad_sumar = request.POST.get('suma', '0').replace(',', '.')
        cantidad_sumar = Decimal(cantidad_sumar)
        
        if cantidad_sumar < 0:
            raise error_contexto("La cantidad a sumar no puede ser negativa.")
    except (ValueError, TypeError):
        raise error_contexto("Cantidad inválida.")

    alimento.cantidad += cantidad_sumar
    alimento.save()

    url_redireccion = request.session.get('url_panel_principal', reverse('panel_principal'))
    return HttpResponseRedirect(url_redireccion)


#@manejo_errores
@require_POST
def restar_cantidad_alimento(request, id_alimento):
    alimento = get_object_or_404(Alimento, pk=id_alimento)

    try:
        cantidad_restar = request.POST.get('resta', '0').replace(',', '.')
        cantidad_restar = Decimal(cantidad_restar)
    
        if cantidad_restar < 0:
            raise error_contexto("La cantidad a restar no puede ser negativa.")
        
    except (ValueError, TypeError):
        raise error_contexto("Cantidad inválida.")

    nueva_cantidad = max(alimento.cantidad - cantidad_restar, Decimal('0'))
    alimento.cantidad = nueva_cantidad
    alimento.save()

    url_redireccion = request.session.get('url_panel_principal', reverse('panel_principal'))
    return HttpResponseRedirect(url_redireccion)




"""===> OTRAS VISTAS <==="""
@manejo_errores
@require_safe
def generar_excel(request):
    filtro_desc_alimento = request.GET.get('filtro_desc_alimento', '').strip()
    filtro_almacen_id = request.GET.get('filtro_almacen', '').strip()

    alimentos_qs = Alimento.objects.filter(almacen_perteneciente__estado=True)

    if filtro_desc_alimento:
        alimentos_qs = alimentos_qs.filter(desc_alimento__icontains=filtro_desc_alimento)

    if filtro_almacen_id.isdigit():
        alimentos_qs = alimentos_qs.filter(almacen_perteneciente_id=int(filtro_almacen_id))

    alimentos_qs = alimentos_qs.order_by('id_alimento')

    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Descripción Alimento'
    ws['B1'] = 'Almacén'
    ws['C1'] = 'Cantidad (gramos)'

    row = 2
    for alimento in alimentos_qs:
        ws[f'A{row}'] = alimento.desc_alimento
        ws[f'B{row}'] = alimento.almacen_perteneciente.desc_almacen if alimento.almacen_perteneciente else "Sin almacén"
        ws[f'C{row}'] = float(alimento.cantidad)
        row += 1

    nombre_archivo = "Reporte_Alimentos.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    wb.save(response)
    return response




